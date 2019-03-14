# -*- coding: utf-8 -*-
author_= '新来的小学生'
import json
import re
import time
import urllib.parse
from random import randint
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import Workbook
from datetime import datetime
# from 优衣库爬虫.db import tieba_collection

options = webdriver.ChromeOptions()
driver = webdriver.Chrome(chrome_options=options)

Headers = {
    'Accept':'application/json, text/plain, */*',
    'Accept-Encoding':'gzip, deflate, br',
    'Host':'tieba.baidu.com',
    'User-Agent':'Mozilla/5.0 (Linux; Android 7.0; KNT-UL10 Build/HUAWEIKNT-UL10) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/37.0.0.0 Mobile Safari/537.36'
    }

#判断下一页是否存在
def tie_next_page():
    try:
        for element in driver.find_elements_by_xpath('//ul[@class="l_posts_num"]'):
            next_page = element.find_element_by_link_text('下一页')
            action = ActionChains(driver)
            action.move_to_element(next_page).perform()
            time.sleep(randint(1,2))
            next_page.click()
            return True
        print('no next page')
        return False
    except Exception as e:
        driver.refresh()
        return False

#获取全吧搜索下所有帖子信息
def get_tie_list(url):
    tie_list=[]
    for p in range(1,77):
        tieurl = url.format(i= p)
        print(tieurl)
        try:
            htm = requests.get(tieurl).text
            time.sleep(1)
            sou = BeautifulSoup(htm, 'lxml')
            conten = sou.text
            jsondat = json.loads(conten)
            dat = jsondat['data']['data']
        except:
            continue
        try:
            datalist = dat['post']
        except:
            continue
        for da in datalist:
            tiedict = {}
            title = re.sub(r'回复:','',da['title'])
            title = re.sub(r'<\w*>','',title)
            title = re.sub(r'</\w*>', '', title)
            tiedict['title'] = title
            tiedict['tid'] = da['tid']
            tiedict['link'] = da['url']
            tiedict['forum'] = da['forum']
            tiedict['createdtime'] = da['time']
            tie_list.append(tiedict)
            #全吧搜索下所有帖子的概述信息 [{'tid': '5957540317', 'title': '对象想在优衣库里上了我', 'createdtime': '1542872788', 'link': 'http://tieba.baidu.com/f?kz=5957540317&pid=122957242123', 'forum': 'les'}, {'tid': '5957360876', 'title': '震惊 震惊!又一品牌辱华被封杀', 'createdtime': '1542872296', 'link': 'http://tieba.baidu.com/f?kz=5957360876&pid=122957154912', 'forum': 'wp7'}]
    final_tie_list = []

    #去重
    for tie in tie_list:
        if tie['tid'] not in final_tie_list:
            final_tie_list.append(tie['tid'])
            final_tie_list.append(tie)
    tielist = []
    for f_tie in final_tie_list:
        if type(f_tie)== dict:
            tielist.append(f_tie)
            with open('帖子id集合.txt','a+',encoding='utf-8') as f:
                f.write(str(f_tie)+'\n')
    return tielist  #获取最终符合条件的帖子概述信息集合


#获取页面下的所有楼层信息
def get_tie_data_field(forum_id, thread_id, page_no):
    time.sleep(randint(1,2))
    for element in driver.find_elements_by_xpath('//div[@class="p_postlist"]/div'):  #某页下的所有楼层
        try:
            ele_data_field = element.get_attribute('data-field')
            if ele_data_field == None:
                continue
            #对接口返回的json数据处理修改再将结果打包存储
            content_data_field = json.loads(ele_data_field)
            if content_data_field.get('content') == None: #判断该楼层是否有内容  无内容跳过此楼  （判断广告）
                continue
            if content_data_field['content'].get('forum_id') == None:   #想获取此楼层的所有评论信息 需要吧id 贴id 唯一pid  构建api请求  若楼层信息没带 将信息給予它
                content_data_field['content']['forum_id'] = forum_id
            if content_data_field['content'].get('thread_id') == None:
                content_data_field['content']['thread_id'] = thread_id
            if forum_id != content_data_field['content'].get('forum_id'):
                continue
            if thread_id != content_data_field['content'].get('thread_id'):
                continue
            if content_data_field['content'].get('content') == None:
                _post_id = content_data_field['content'].get('post_id')  #链接评论信息唯一标识
                content = element.find_element_by_id('post_content_'+str(_post_id)).text  #楼层回复内容
                content_data_field['content']['content'] = content
            else:
                content = content_data_field['content']['content']
            try:
                try:
                    content_data_field['content']['date'] =element.find_elements_by_class_name('tail-info')[-1].text
                except:
                    content_data_field['content']['date'] =element.find_element_by_class_name('p_tail').find_elements_by_tag_name('li')[-1].text
            except:
                content_data_field['content']['date'] =''
                #对评论内容进行数据格式处理
            content_text = re.compile('<[^>]+>').sub("", content)
            content_img = re.findall('<[^>]+>', content)
            content_data_field['content']['content_text'] = content_text
            content_data_field['content']['content_img'] = content_img
        except:
            continue
        tie_data_field.append({'author': content_data_field['author'], 'content': content_data_field['content']})
    # 查看是否有下一页
    flag = tie_next_page()
    if flag:
        page_no = page_no + 1
        return get_tie_data_field(forum_id, thread_id, page_no)
    return page_no


#获得此帖子所有页面下的评论信息
def get_tie_total_comment(forum_id, thread_id, total_page):
    i = 0
    comment_list = {}
    user_list = {}
    while i < total_page:
        i = i + 1
        try:
            comment_url = 'https://tieba.baidu.com/p/totalComment?tid={tid}&fid={fid}&pn={pn}&see_lz=0'.format(tid=thread_id, fid=forum_id, pn=i)
            rsp = _session.get(comment_url)
            time.sleep(1)
            data = json.loads(rsp.text)['data']
            if len(data['comment_list']) == 0:
                continue
            if len(data['user_list']) == None:
                continue
            comment_list.update(data['comment_list'])
            user_list.update(data['user_list'])
        except Exception as e:
            print('comment error')
            break
    return {'comment_list': comment_list, 'user_list': user_list}





def write_tie_excel(list):
    wb = Workbook()
    ws=wb.active
    ws['A1'] = '标题'
    ws['B1'] = '帖子ID'
    ws['C1'] = 'url地址'
    ws['D1'] = '吧名'
    ws['E1'] = '帖子创建时间'
    ws['F1']='帖子评论数'
    ws['G1']='昵称'
    ws['H1']='评论内容'
    ws['I1']='评论时间'
    ws['J1']='回复数'
    ws['K1']='昵称'
    ws['L1']='回复时间'
    ws['M1']='回复内容'
    row=2
    for m in list:
        if m:
            ws.cell(row=row,column=1,value=m['title'])
            ws.cell(row=row,column=2,value=m['id'])
            ws.cell(row=row,column=3,value=m['url'])
            ws.cell(row=row,column=4,value=m['forum'])
            ws.cell(row=row,column=5,value=m['createdtime'])
            ws.cell(row=row,column=6,value=m['reply_number'])
            level_1_content=m['content']
            for level2 in level_1_content:
                col=7
                values=[level2['author_name'],level2['tie_content'],level2['creat_time'],level2['tie_content_number']]
                for i in values:
                    ws.cell(row=row+1,column=col,value=i)
                    col=col+1
                level_2_content=level2['content']
                if level_2_content:
                    for level3 in level_2_content:
                        col=11
                        row=row+1
                        values3=[level3['name'],level3['time'],level3['content']]
                        for j in values3:
                            ws.cell(row=row+1,column=col,value=j)
                            col=col+1
                else:
                    pass
                row=row+1
            row=row+1
        else:
            continue
    now=datetime.now().date()
    filename=str(now)+"优衣库.xlsx"
    wb.save(filename)




if __name__ == '__main__':

    username_ = '15555721298'
    password_ = 'su521125211314'
    tieurl = 'https://tieba.baidu.com/index.html'
    driver.get(tieurl)
    try:
        login_btn = driver.find_element_by_xpath('//*[@id="com_userbar"]/ul/li[4]/div/a')
        login_btn.click()
        time.sleep(1)
        user_login=driver.find_element_by_xpath('//*[@id="TANGRAM__PSP_10__footerULoginBtn"]')
        user_login.click()
        time.sleep(1)
        username = driver.find_element_by_xpath('//*[@id="TANGRAM__PSP_10__userName"]')
        password = driver.find_element_by_xpath('//*[@id="TANGRAM__PSP_10__password"]')
        time.sleep(1)
        username.clear()
        username.send_keys(username_)
        password.clear()
        password.send_keys(password_)
        time.sleep(1)
        sub_btn = driver.find_element_by_xpath('//*[@id="TANGRAM__PSP_10__submit"]')
        sub_btn.click()
        time.sleep(30)
    except:
        print('自动登录失败')
        time.sleep(5)

    _session = requests.session()
    def init_req_session():
        chrome_cookies = driver.get_cookies()
        cookies = {}
        for c in chrome_cookies:
            cookies[c['name']] = c['value']
        requests.utils.add_dict_to_cookiejar(_session.cookies, cookies)


    keys = 'Uniqlo'
    # keys = '优衣库'

    values = {}
    values['word'] = keys
    data = urllib.parse.urlencode(values)
    url = "https://tieba.baidu.com/mo/q/seekcomposite?pn={i}&rn=10&is_ajax=1&sort=1&"+data  # 优衣库
    tielist = []
    tielist = get_tie_list(url)
    idlist=[]
    #获取总帖子
    # with open('帖子id集合.txt','rb') as f:
    #     lines=f.readlines()
    # for line in lines:
    #     dict=eval(line)
    #     tielist.append(dict)
    list_len=len(tielist)
    num=1
    totaldatalist=[]
    for tie in tielist:
        level_1={}
        tie_data_field = []
        if tie['tid'] not in idlist:
            try:
                level_2_list=[]
                tie_content = []
                print('开始爬取第'+str(num)+'个帖子 共'+str(list_len))
                url = tie['link']
                level_1['title']=tie['title'] #帖子标题
                level_1['url']=url  #帖子链接
                level_1['id']=tie['tid'] #帖子id
                level_1['forum']=tie['forum'] #吧名
                # if tie['createdtime']:
                #     timeStamp = int(tie['createdtime']) #帖子最新回复时间戳
                #     timeArray = time.localtime(timeStamp)
                #     level_1['createdtime']=  time.strftime("%Y-%m-%d %H:%M:%S", timeArray)   #帖子创建时间
                #     #格式化时间数据
                #     tie_timedate=str(level_1['createdtime']).split(' ')[0]
                #     tie_time=datetime.strptime(tie_timedate, "%Y-%m-%d")
                #     begin_date='2018-10-25'
                #     begin_date = datetime.strptime(begin_date, "%Y-%m-%d")
                #     end_date='2018-11-25'
                #     end_date = datetime.strptime(end_date, "%Y-%m-%d")
                #     if begin_date<=tie_time<=end_date:
                #         pass
                #     else:
                #         continue
                # else:
                #     level_1['createdtime']=''
                try:
                    driver.get(url)
                except:
                    continue
                try:
                    forum_id = int(re.findall(r'forum_id:\s+"(\d+)",', driver.page_source, re.M)[0])      #获取吧 id
                    thread_id = int(re.findall(r'thread_id:(\d+),', driver.page_source, re.M)[0])   #kz值   #获取帖子 id
                except:
                    forum_id = int(re.findall(r'"forum_id":(\d+),', driver.page_source, re.M)[0])
                    thread_id = int(re.findall(r'"thread_id":"(\d+)",', driver.page_source, re.M)[0])
                time.sleep(randint(1,2))
                html=driver.page_source
                soup=BeautifulSoup(html,'lxml')
                try:
                    level_1['reply_number']=soup.find_all(class_='red')[0].text #帖子评论数
                except:
                    level_1['reply_number']=''
         #获取一楼帖子的回复时间  即帖子的创建时间
                try:
                    try:
                        element=soup.find_all(class_="j_l_post")[0]
                        level_1['createdtime'] =element.find_all(class_='tail-info')[-1].text
                    except:
                        element=soup.find_all(class_="p_postlist")[0]
                        level_1['createdtime'] =element.find(class_='p_tail').find_all(name='li')[-1].text
                except:
                    level_1['createdtime'] =''
                #判断是否跳出
                if level_1['createdtime']:
                    #格式化时间数据
                    tie_timedate=str(level_1['createdtime']).split(' ')[0]
                    tie_time=datetime.strptime(tie_timedate, "%Y-%m-%d")
                    begin_date='2018-10-25'
                    begin_date = datetime.strptime(begin_date, "%Y-%m-%d")
                    end_date='2018-11-25'
                    end_date = datetime.strptime(end_date, "%Y-%m-%d")
                    if begin_date<=tie_time<=end_date:
                        pass
                    else:
                        num+=1
                        with open('已爬取的帖子id集合.txt','a+',encoding='utf-8') as f:
                            f.write(level_1['id']+'\n')
                        continue

                total_page = get_tie_data_field(forum_id, thread_id, 1)   #获取此帖子下页数 与 tie_data_field 存储帖子下所有楼层回复信息
                if total_page>300:
                    total_page = 300


                total_comment = get_tie_total_comment(forum_id, thread_id, total_page)   #此帖子下所有页面 楼层下评论信息的字典集合


                for index,tie_data in enumerate(tie_data_field):
                    print("楼层"+str(index+1),len(tie_data_field))
                    #判断此楼层下是否有评论数
                    level_2={}
                    level_2['author_name']=tie_data['author']['user_name']  #楼层回复人昵称
                    level_2['creat_time']=tie_data['content']['date']  #楼层回复时间
                    level_2['tie_content']=tie_data['content']['content_text']   #楼层回复内容
                    #判断该楼层是否有评论
                    if tie_data['content']['comment_num'] > 0:
                        level_2['tie_content_number']=tie_data['content']['comment_num']  #楼层下的评论数
                        post_id = tie_data['content']['post_id']
                        comment = total_comment['comment_list'][str(post_id)]  #此post_id 标识的该楼层下所有评论的字典集合
                        tie_data['comment'] = comment.copy()
                        level_3_list=[]
                        for c_info in tie_data['comment']['comment_info']:  #楼层下的所有评论信息列表
                            level_3={}
                            level_3['name']=c_info['username']  #评论人用户名
                            content = c_info['content']
                            content_text = re.compile('<[^>]+>').sub("", content)
                            content_text = re.compile('回复[^.]+[:|：]').sub("", content_text)
                            # content_img = re.findall('<[^>]+>', content)
                            level_3['content']= content_text  #评论内容
                            # c_info['content_img'] = content_img
                            commenttime=c_info['now_time']  #评论时间
                            commenttimeArray = time.localtime(commenttime)
                            level_3['time']=  time.strftime("%Y-%m-%d %H:%M:%S", commenttimeArray)
                            level_3_list.append(level_3)
                        level_2['content']=level_3_list
                    else:
                        level_2['tie_content_number']='0'  #楼层回复下的评论数
                        level_2['content']='' #楼层回复下的评论内容
                    level_2_list.append(level_2)
                level_1['content']=level_2_list
                totaldatalist.append(level_1)

                with open('临时贴吧.txt','a+',encoding='utf-8') as f:
                    f.write(str(level_1)+'\n')
            except:
                continue
            num+=1
            with open('已爬取的帖子id集合.txt','a+',encoding='utf-8') as f:
                f.write(level_1['id']+'\n')
        else:
            num+=1
            continue
    write_tie_excel(totaldatalist)
    driver.close()


