from openpyxl import Workbook


def write_tie_excel(list,key):
    wb = Workbook()
    ws=wb.active
    ws['A1'] = '标题'
    ws['B1'] = '帖子ID'
    ws['C1'] = 'url地址'
    ws['D1'] = '吧名'
    ws['E1'] = '帖子创建时间'
    ws['F1']='帖子评论数'
    row=2
    col=1
    for m in list:
        values=[m['title'],m['id'],m['url'],m['ba_name'],m['tie_creattime'],m['reply_number']]
        for i in values:
            ws.cell(row=row,column=col,value=i)
            col=col+1
        row=row+1
        col=1
    filename=key+'-'+"帖子信息.xlsx"
    wb.save(filename)


if __name__ == '__main__':
    list=[]
    with open('尸兄帖子信息.txt','r',encoding='utf-8') as f:
        lines=f.readlines()
        for line in lines:
            dict=eval(line)
            list.append(dict)
    write_tie_excel(list,'尸兄')