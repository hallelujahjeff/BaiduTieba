# -*- coding: utf-8 -*-
# time=2018.11.21
"""
info:
author : 'su'
github:https://github.com/suxin1995/
update_time:2019-4-30
"""


import requests
from random import randint
import time
from bs4 import BeautifulSoup
import json
import re
from openpyxl import Workbook
from datetime import datetime
from urllib import parse
from fake_useragent import UserAgent


class Tieba_Spider:

    def __init__(self, keyword, max_page):
        self.keyword = parse.quote(keyword)
        self.max_page = max_page
        self.base_url = 'https://tieba.baidu.com/f?kw={}&ie=utf-8&pn={}'
        self.Headers = {
            'Accept':'application/json, text/plain, */*',
            'Accept-Encoding':'gzip, deflate, br',
            'Host':'tieba.baidu.com',
            'User-Agent':'Mozilla/5.0 (Linux; Android 7.0; KNT-UL10 Build/HUAWEIKNT-UL10) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/37.0.0.0 Mobile Safari/537.36'
        }

    #获取url列表
    def get_url_list(self):
        return  [self.base_url.format(self.keyword, pn) for pn in range(0, self.max_page, 50)]

    #获取网页内容
    def get_content(self,url):
        response = requests.get(url,headers = self.Headers)
        content = response.text
        code = response.status_code
        return content,code

    #解析网页内容
    def parse_content(self,content):
        tieid_list = []
        soup = BeautifulSoup(content, 'html.parser')
        tie_list = soup.find_all(class_='tl_shadow')
        for m in tie_list:
            try:
                tid = m.find(class_='j_common').attrs['data-tid']
                tieid_list.append(tid)
            except:
                continue
        return tieid_list

    #解析帖子内容
    def get_tie(self,id):
        datalist = []
        tiedatalist = []
        for page in range(0, 50):
            url = 'https://tieba.baidu.com/p/' + str(id) + '?pn=' + str(page)
            try:
                html = requests.get(url,headers = self.Headers).text
                time.sleep(randint(1, 2))
                soup = BeautifulSoup(html, 'html.parser')
                title = soup.find(name='title').text #帖子标题
                tie_url =  'https://tieba.baidu.com/p/' + str(id)  #帖子url
                try:
                    amount = soup.find(class_='l_reply_num').find_all(class_='red')[-1].text
                except:
                    amount = 10
                try:
                    dict['reply_number'] = soup.find(class_='l_reply_num').find_all(class_='red')[-2].text  # 帖子总回复数
                except:
                    dict['reply_number'] = ''
                try:
                    content_list = soup.find_all(class_='l_post')
                    if i == 1:  # 检测一次
                        try:
                            try:
                                dict['ba_name'] = soup.find(class_='card_title_fname').text  # 吧名
                            except:
                                dict['ba_name'] = ''
                            try:
                                dict['title'] = soup.find(class_='core_title_txt').attrs['title']  # 发帖主题
                            except:
                                dict['title'] = ''
                            tietime = content_list[0].find_all(class_='tail-info')[-1].text  # 发帖时间
                            dict['tie_creattime'] = tietime
                            date = tietime.split(' ')[0]
                            tie_time = datetime.strptime(date, "%Y-%m-%d")
                            begin_date = '2018-1-1'
                            begin_date = datetime.strptime(begin_date, "%Y-%m-%d")
                            end_date = '2018-10-30'
                            end_date = datetime.strptime(end_date, "%Y-%m-%d")
                            if begin_date <= tie_time <= end_date:
                                pass
                            else:
                                break

                        except:
                            dict['tie_creattime'] = ''
                except:
                    print('获取帖子错误：%s'%url)
                    break

                i += 1

                for i in content_list:
                    tiedata = {}
                    try:
                        tiedata['author_name'] = i.find(class_='d_name').text  # 用户名
                    except:
                        tiedata['author_name'] = ''
                    try:
                        tiedata['creat_time'] = i.find_all(class_='tail-info')[-1].text  # 回复时间
                    except:
                        tiedata['creat_time'] = ''
                    try:
                        tiedata['tie_content'] = i.find(class_='d_post_content').text  # 回复内容
                    except:
                        tiedata['tie_content'] = ''
                    # tie_list.append(tiedata)
            except:
                break
            try:
                if page == int(amount) - 1:
                    break
            except:
                pass
        # if tie_list:
        #     option['content'] = tie_list  # 帖子内容
        #     option['id'] = m  # 帖子id
        #     with open('帖子回复.txt', 'a+', encoding='utf-8') as f:
        #         f.write(str(option) + '\n')  # 记录临时每个帖子的信息
        #     datalist.append(option)
        # else:
        #     continue
        # dict['title'] = option['title']  # 帖子主题
        # dict['id'] = m  # 帖子id
        # dict['url'] = 'https://tieba.baidu.com/p/' + str(m)  # 帖子url
        # with open('帖子信息.txt', 'a+', encoding='utf-8') as f:
        #     f.write(str(dict) + '\n')
        # tiedatalist.append(dict)

    #保存格式化数据
    def save_content(self):
        pass







class ProcessData:

    def __int__(self, flag, list, key):
        self.flag = flag
        self.list = list
        self.key = key

    def write_information_excel(self):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '帖子ID'
        ws['B1'] = '帖子标题'
        ws['C1'] = '用户名称'
        ws['D1'] = '回复内容'
        ws['E1'] = '回复时间'
        row = 2
        col = 1
        for m in self.list:
            content = m['content']
            for n in content:
                values = [m['id'], m['title'], n['author_name'], n['tie_content'], n['creat_time']]
                for i in values:
                    ws.cell(row=row, column=col, value=i)
                    col = col + 1
                row = row + 1
                col = 1
        filename = self.key + '-' + "reply.xlsx"
        wb.save(filename)

    def write_tie_excel(self):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '标题'
        ws['B1'] = '帖子ID'
        ws['C1'] = 'url地址'
        ws['D1'] = '吧名'
        ws['E1'] = '帖子创建时间'
        ws['F1'] = '帖子评论数'
        row = 2
        col = 1
        for m in self.list:
            values = [m['title'], m['id'], m['url'], m['ba_name'], m['tie_creattime'], m['reply_number']]
            for i in values:
                ws.cell(row=row, column=col, value=i)
                col = col + 1
            row = row + 1
            col = 1
        filename = self.key + '-' + "post.xlsx"
        wb.save(filename)
    # write_tie_excel(tiedatalist, key)
    # write_information_excel(datalist, key)





if __name__ == '__main__':
    key = 'python'  # 吧名
    max_page = 2
    Tie_list = []
    tieba = Tieba_Spider(key,max_page)
    #获取整个吧所有页面URL
    url_list = tieba.get_url_list()
    for url in url_list:
        html,code = tieba.get_content(url)
        if code == 200:
            #获取每个页面所有帖子URL
            tieid_list = tieba.parse_content(html)
            for n in tieid_list:
                if n not in Tie_list:
                    Tie_list.append(n)
        else:
            print('错误URL:',url)
    print(Tie_list)
    for url in Tie_list:
        tieba.get_tie(url)


