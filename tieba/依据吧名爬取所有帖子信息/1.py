from openpyxl import Workbook

def write_information_excel(list,key):
    wb = Workbook()
    ws=wb.active
    ws['A1'] = '帖子ID'
    ws['B1'] = '帖子标题'
    ws['C1'] = '用户名称'
    ws['D1'] = '发帖时间'
    ws['E1'] = '帖子内容'
    row=2
    col=1
    for m in list:
        values=[m['id'],m['title'],m['author_name'],m['tietime'],m['tie_content']]
        for i in values:
            ws.cell(row=row,column=col,value=i)
            col=col+1
        row=row+1
        col=1
    filename=key+'-'+"数据.xlsx"
    wb.save(filename)

with open('师大帖子.txt','r',encoding='utf8') as f:
    lines=f.readlines()
datalist=[]
for line in lines:
    line=eval(line)
    datalist.append(line)
write_information_excel(datalist,'南京师范大学')