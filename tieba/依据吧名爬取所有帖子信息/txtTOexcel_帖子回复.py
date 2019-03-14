from openpyxl import Workbook


def write_information_excel(list,key):
    wb = Workbook()
    ws=wb.active
    ws['A1'] = '帖子ID'
    ws['B1'] = '帖子标题'
    ws['C1'] = '用户名称'
    ws['D1'] = '回复内容'
    ws['E1'] = '回复时间'
    row=2
    col=1
    for m in list:
        content=m['content']
        for n in content:
            values=[m['id'],m['title'],n['author_name'],n['tie_content'],n['creat_time']]
            for i in values:
                ws.cell(row=row,column=col,value=i)
                col=col+1
            row=row+1
            col=1
    filename=key+'-'+"帖子回复.xlsx"
    wb.save(filename)


if __name__ == '__main__':
    list=[]
    with open('尸兄帖子回复.txt','r',encoding='utf-8') as f:
        lines=f.readlines()
        for line in lines:
            dict=eval(line)
            list.append(dict)
    write_information_excel(list,'尸兄')